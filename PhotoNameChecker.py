import streamlit as st
import os
import unicodedata
import re
import requests
from bs4 import BeautifulSoup
import pandas as pd
from urllib.parse import urlparse, parse_qs
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from io import BytesIO
import time

# --- Common HTTP Headers to avoid basic bot detection and connection issues ---
COMMON_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Connection': 'keep-alive'
}

# --- Filename parsing (Unchanged) ---

def parse_filenames(folder_or_files):
    """
    Parse filenames in the format:
      school_last_first[_anything].png
    Ignores any extra suffixes after first name.
    """
    parsed = []
    for file in folder_or_files:
        name = os.path.basename(file)
        if not name.lower().endswith(".png"):
            continue

        base = name[:-4]  # strip .png
        parts = base.split("_")

        if len(parts) < 3:
            parsed.append({
                "filename": name,
                "school": None,
                "last": None,
                "first": None,
                "format_valid": False,
                "format_msg": "Too few parts: need school_last_first"
            })
            continue

        school, last, first = parts[0], parts[1], parts[2]

        parsed.append({
            "filename": name,
            "school": school.lower().strip(),
            "last": last.lower().strip(),
            "first": first.lower().strip(),
            "format_valid": True,
            "format_msg": None
        })

    return parsed

# --- Normalization / roster scraping ---

def normalize(name: str) -> str:
    """
    Normalize names to your conventions, including removing nicknames.
    Handles nicknames in single or double quotes.
    """
    name = name.lower()
    # Remove nicknames in single or double quotes: "nickname" or 'nickname'
    name = re.sub(r'(["“”‘’\']).*?\1', '', name)
    # Remove suffixes like Jr, Sr, II, III
    name = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", name)
    # Remove accents
    name = unicodedata.normalize("NFD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    # Remove non-word characters except spaces/hyphens
    name = re.sub(r"[^\w\s-]", "", name)
    # Collapse multiple spaces and trim
    name = re.sub(r"\s+", " ", name).strip()
    return name

# --- Baylor scraping functions (use Selenium, no change needed) ---

def scrape_baylor_players(url: str):
    import requests
    from bs4 import BeautifulSoup

    players = {}
    nickname_names = {}

    r = requests.get(url, headers=COMMON_HEADERS, timeout=30, verify=False)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    for a in soup.select("div[data-test-id='s-person-details__personal-single-line'] a"):
        href = a.get("href") or ""
        if "/coaches/" in href:
            continue
        h3 = a.find("h3")
        if h3:
            name = h3.get_text(strip=True)
            players[normalize(name)] = name
    return players, nickname_names
    
def scrape_baylor_staff(url: str):
    """
    Scrape Baylor staff/coaches using Requests + BeautifulSoup (no Selenium).
    Returns a dictionary: {normalized_name: {"name": original_name, "title": "Coach"}}
    """
    import requests
    from bs4 import BeautifulSoup

    staff_dict = {}

    try:
        r = requests.get(url, headers=COMMON_HEADERS, timeout=30, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # Select all the 'personal single line' divs
        for div in soup.select("div[data-test-id='s-person-details__personal-single-line'] a"):
            href = div.get("href") or ""
            h3 = div.find("h3")
            if not h3:
                continue
            name = h3.get_text(strip=True)

            # Only pick coaches/staff (links with '/coaches/')
            if "/coaches/" in href:
                staff_dict[normalize(name)] = {"name": name, "title": "Coach"}

        return staff_dict

    except Exception as e:
        st.error(f"Error scraping Baylor staff (BS version): {e}")
        return {}


def contains_invalid_word(name: str, invalid_words: list[str]) -> bool:
    """
    Returns True if any invalid word matches as a whole word in the name.
    """
    name_lower = name.lower()
    for word in invalid_words:
        if re.search(rf"\b{re.escape(word)}\b", name_lower):
            return True
    return False

def clean_roster_name(raw_name: str) -> str:
    """
    Ensure roster names are consistently 'First Last'
    even if site lists them as 'Last, First'.
    """
    raw_name = raw_name.strip()
    if "," in raw_name:  # handles "Alexander, Kael"
        parts = [p.strip() for p in raw_name.split(",", 1)]
        if len(parts) == 2:
            last, first = parts
            return f"{first} {last}"
    return raw_name

def fix_character_confusion(name: str) -> str:
    """
    Fix common OCR/font character confusions in scraped names.
    Example: "Mclntosh" -> "McIntosh"
    """
    # Fix Mc[lowercase L] followed by consonants -> Mc[uppercase I]
    # This handles "Mclntosh" -> "McIntosh"
    name = re.sub(r'\bMcl([bcdfghjkmnpqrstvwxyz])', r'McI\1', name, flags=re.IGNORECASE)
    
    return name

def is_valid_player_name(name: str) -> bool:
    """
    Check if a name is valid for a player (has at least first and last name).
    Returns False for single-word names or obviously invalid entries.
    """
    if not name or len(name.strip()) == 0:
        return False
    
    # Split by spaces
    parts = name.strip().split()
    
    # Must have at least 2 words (first and last name)
    if len(parts) < 2:
        return False
    
    # Each part should have at least 2 characters (excludes initials-only entries)
    # Exception: Allow single letters if they're followed by a period (like "T.J.")
    for part in parts:
        clean_part = part.replace('.', '')
        if len(clean_part) == 0:
            return False
    
    return True

def remove_credentials(name: str) -> str:
    """
    Remove common credentials/titles from names.
    Examples: MS, ATC, LAT, PhD, MD, MBA, etc.
    Also removes nicknames in quotes.
    """
    # Remove anything between quotes (handles all quote types)
    # Use Unicode escape codes for curly quotes to avoid syntax errors
    
    # Curly double quotes: " " (U+201C and U+201D)
    name = re.sub(r'[\u201c\u201d][^\u201c\u201d]*[\u201c\u201d]', '', name)
    # Straight double quotes: " "
    name = re.sub(r'"[^"]*"', '', name)
    # Curly single quotes: ' ' (U+2018 and U+2019)
    name = re.sub(r'[\u2018\u2019][^\u2018\u2019]*[\u2018\u2019]', '', name)
    # Straight single quotes: ' '
    name = re.sub(r"'[^']*'", '', name)
    
    # Remove credentials pattern like ", MS, ATC, LAT"
    name = re.sub(r',\s*[A-Z]{2,}(\s*,\s*[A-Z]{2,})*', '', name)
    # Remove trailing commas
    name = re.sub(r',\s*$', '', name).strip()
    # Clean up extra spaces
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def scrape_wyoming_players_selenium(url: str):
    """
    Scrape Wyoming players using Selenium (JavaScript-rendered page).
    """
    found_names = set()
    
    try:
        # Setup Chrome options
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        
        # Detect if running on Streamlit Cloud or similar Linux environment
        try:
            # Try to use system chromium-driver (for Streamlit Cloud)
            driver = webdriver.Chrome(options=chrome_options)
        except:
            # Fallback to ChromeDriverManager (for local development)
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        
        driver.get(url)
        
        # Wait for JavaScript to load the roster
        time.sleep(5)
        
        # Find all roster links
        elements = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/roster/"], div.sidearm-roster-player-name a')
        
        for elem in elements:
            try:
                href = elem.get_attribute('href') or ''
                
                # Skip coaches/staff - they have /coaches/ or /staff/ in the URL
                if '/coaches/' in href or '/staff/' in href:
                    continue
                
                name = elem.get_attribute('title') or elem.text
                name = name.strip()
                
                if name:
                    # Clean up extra spaces
                    name = re.sub(r'\s+', ' ', name)
                    found_names.add(name)
            except:
                continue
        
        driver.quit()
        return found_names
        
    except Exception as e:
        st.error(f"Selenium error for Wyoming: {e}")
        return set()

def scrape_player_names(url: str):
    """
    Scrape player names from a roster page.
    Returns two dictionaries:
        primary_names: {normalized_name: original_name}
        nickname_names: {normalized_name: original_name} (if nicknames detected)
    """
    is_baylor = "baylorbears.com" in url.lower()
    is_stetson = "stetson.edu" in url.lower()
    is_wyoming = "gowyo.com" in url.lower() or "wyoming" in url.lower()
    is_george_mason = "gomason.com" in url.lower()
    is_siena = "sienasaints.com" in url.lower()  # ADD THIS
    
    found_names = set()
    invalid_keywords = [
        "news", "schedule", "statistics", "videos",
        "links", "gameday", "staff", "coach", "bio", "media",
        "ireland", "tarheels2ireland", "central", "additional",
        "more", "results", "events", "©", "menu", "25fb", "2025",
        "photo", "headshot", "print", "roster", "search", "jump",
        "video", "feb", "mar", "stats", "baseball",
        "donate", "coaches", "camps"
    ]
    
    try:
        resp = requests.get(url, timeout=30, headers=COMMON_HEADERS, verify=False)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        
        # Wyoming and George Mason specific scraping (JavaScript-rendered)
        if is_wyoming or is_george_mason:
            found_names = scrape_wyoming_players_selenium(url)
            if found_names:
                primary_names = {normalize(name): name for name in found_names}
                return primary_names, {}
            else:
                st.warning("No Wyoming players found with Selenium")
                return {}, {}
        
        # Baylor logic
        if is_baylor:
            return scrape_baylor_players(url)
        
        # Stetson-specific scraping
        if is_stetson:
            for li in soup.select("li.sidearm-roster-player"):
                name_tag = li.select_one("h3 a")
                if name_tag:
                    raw_name = name_tag.get_text(" ", strip=True)
                    name = clean_roster_name(raw_name)
                    
                    # Use the new validation function
                    if name and is_valid_player_name(name) and not contains_invalid_word(name, ["news", "schedule", "staff", "coach", "video"]):
                        found_names.add(name)
            
            primary_names = {normalize(name): name for name in found_names}
            nickname_names = {}
            return primary_names, nickname_names
        
        # --- Siena-specific scraping (ADD THIS BLOCK) ---
        if is_siena:
            # Siena uses a specific structure - target player links directly
            for a_tag in soup.select('li.sidearm-roster-player a[href*="/roster/"]'):
                # Skip staff links
                if '/staff/' in a_tag.get('href', ''):
                    continue
                    
                # Try to find name in h3 tag
                h3_tag = a_tag.find('h3')
                if h3_tag:
                    name = h3_tag.get_text(" ", strip=True)
                    name = clean_roster_name(name)
                    if name and not contains_invalid_word(name, invalid_keywords):
                        found_names.add(name)
            
            # If we found names, return them
            if found_names:
                st.info(f"Siena scraper found {len(found_names)} players")  # DEBUG
                primary_names = {normalize(name): name for name in found_names}
                return primary_names, {}
            
        # --- GENERIC URL-based player/staff separation ---
        # Try to detect if this site uses /player/ and /staff/ URL patterns
        player_links = soup.select('a[href*="/player/"]')
        staff_links = soup.select('a[href*="/staff/"], a[href*="/coaches/"]')
            
        # If we found player links with /player/ pattern, use URL-based filtering
        if len(player_links) >= 5:  # At least 5 players to be confident
            #st.info(f"Using URL-based player/staff separation (found {len(player_links)} player links)")
            
            for a_tag in player_links:
                # Double-check it's not a staff link
                href = a_tag.get('href', '')
                if '/staff/' in href or '/coaches/' in href:
                    continue
                
                # Extract name from various possible locations
                span_tag = a_tag.find('span')
                if span_tag:
                    name = span_tag.get_text(" ", strip=True)
                else:
                    name = a_tag.get_text(" ", strip=True)
                
                name = clean_roster_name(name)
                name = fix_character_confusion(name)
                
                # Add validation check here
                if name and is_valid_player_name(name) and not contains_invalid_word(name, invalid_keywords):
                    found_names.add(name)
            
            # If we found names, return them
            if found_names:
                primary_names = {normalize(name): name for name in found_names}
                return primary_names, {}

        # --- Common selectors for other schools ---
        common_player_selectors = [
            ".s-text-regular-bold",
            ".roster-list-item__title",
            ".player-name",
            "td.sidearm-table-player-name",
            "td.sidearm-roster-table-data a[href*='/roster/']",
            ".roster-list-item__name",
            "a.table__roster-name",
            "td.sidearm-roster-table-data a[title]",
            "td > a[href*='/roster/season/']",
            "a.table__roster-name span",
            'div[data-test-id="s-person-details__personal-single-line"] h3',
            'a[href*="/player/"]',
            'li.sidearm-roster-player-name a',
            'div.sidearm-roster-list-item-name a',
            'th.name a[href*="/bios/"]',
            'div.flex-grow-1 a[aria-label]',
            'a[aria-label*="full bio"]',
            'a[aria-label*="View Full Bio"]',
            'a[href*="/roster/"][aria-label]',
            'div[class*="sidearm-roster-list-item-name"] a',
            'a[href*="/roster/"][href^="https://arkansasrazorbacks.com/"]',
            'a[href*="/roster/"]',
            'div.sidearm-roster-list-item-name.sidearm-roster-player-name a',
            'li.sidearm-roster-player h3',  # ADD THIS for Siena
        ]
        
        # --- Step 1: scrape common selectors ---
        for element in soup.select(", ".join(common_player_selectors)):
            name = element.get_text(" ", strip=True)
            name = remove_credentials(name)
            name = clean_roster_name(name)
            name = fix_character_confusion(name)
            
            # Add validation check here
            if name and is_valid_player_name(name) and not contains_invalid_word(name, invalid_keywords):
                found_names.add(name)
        
        # --- Kansas-specific: remove staff from found_names ---
        if "kuathletics.com" in url.lower():
            staff_elements = soup.select('[data-test-id="staff-directory-bio-component__staff-name-value"]')
            for elem in staff_elements:
                staff_name = elem.get_text(strip=True)
                staff_name = clean_roster_name(staff_name)
                if staff_name in found_names:
                    found_names.discard(staff_name)
        
        # --- Step 2: ASU-specific fix: scrape alt from player images ---
        for img_tag in soup.select('a.roster-card__image-wrapper img'):
            name = img_tag.get('alt', '').strip()
            if name:
                found_names.add(name)
        
        # --- Step 3: build primary and nickname dictionaries ---
        primary_names = {}
        nickname_names = {}
        for name in found_names:
            # Look for nicknames in quotes
            match = re.search(r'(\S+)\s+["""''](.+?)["""'']\s+(.+)', name)
            if match:
                first_name, nickname, last_name = match.groups()
                primary_names[normalize(f"{first_name} {last_name}")] = name
                nickname_names[normalize(f"{nickname} {last_name}")] = name
            else:
                normalized = normalize(name)
                
                # If this normalized name already exists, check if one has a suffix
                if normalized in primary_names:
                    existing_name = primary_names[normalized]
                    existing_has_suffix = bool(re.search(r'\b(jr|sr|ii|iii|iv|v)\b', existing_name.lower()))
                    current_has_suffix = bool(re.search(r'\b(jr|sr|ii|iii|iv|v)\b', name.lower()))
                    
                    # Keep the one WITH the suffix (the player, not the coach)
                    if current_has_suffix and not existing_has_suffix:
                        primary_names[normalized] = name
                    # If existing has suffix and current doesn't, keep existing (do nothing)
                    # If both have suffixes or neither do, keep the first one (do nothing)
                else:
                    primary_names[normalized] = name

        return primary_names, nickname_names
        
    except Exception as e:
        st.error(f"Error scraping player names from URL: {e}")
        return {}, {}


def scrape_staff_names(url: str):
    """
    Scrape staff names and titles from a roster page.
    Returns a dictionary: {normalized_name: {"name": original_name, "title": title}}
    """
    staff_dict = {}
    if "baylorbears.com" in url.lower():
        return scrape_baylor_staff(url)
    
    try:
        resp = requests.get(url, timeout=30, headers=COMMON_HEADERS, verify=False)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        
        # --- GENERIC URL-based staff detection ---
        # Try to detect if this site uses /staff/ or /coaches/ URL patterns
        staff_links = soup.select('a[href*="/staff/"], a[href*="/coaches/"]')
        
        # If we found staff links with URL pattern, use URL-based filtering
        if len(staff_links) >= 1:  # At least 1 staff member
            #st.info(f"Using URL-based staff detection (found {len(staff_links)} staff links)")
            
            for a_tag in staff_links:
                # Double-check it's not a player link
                href = a_tag.get('href', '')
                if '/player/' in href:
                    continue
                
                # Extract name
                span_tag = a_tag.find('span')
                if span_tag:
                    name = span_tag.get_text(" ", strip=True)
                else:
                    name = a_tag.get_text(" ", strip=True)
                
                name = remove_credentials(name)
                name = fix_character_confusion(name)
                
                if name:
                    staff_dict[normalize(name)] = {"name": name, "title": "Staff"}
            
            return staff_dict

        # Continue with school-specific checks if generic detection didn't work
        if "thesundevils.com" in url.lower():
            for a_tag in soup.select('h3.roster-card__title a.roster-card__title-link'):
                href = a_tag.get('href', '')
                if '/staff/' in href.lower():
                    name = a_tag.get_text(strip=True)
                    # Get title if available
                    parent_div = a_tag.find_parent('div', class_='roster-card__heading')
                    title = "Staff"
                    if parent_div:
                        info_div = parent_div.find_next_sibling('div', class_='roster-card__info')
                        if info_div:
                            span = info_div.select_one('span.roster-card__position')
                            if span:
                                title = span.get_text(strip=True)
                    staff_dict[normalize(name)] = {"name": name, "title": title}

        # Selectors for various staff list formats
        staff_items = soup.select('li.sidearm-roster-coach, .roster-list-item.staff, tr[data-v-7436a2c8]')
        
        # **FIX 3:** Added robust Sidearm Sports staff selector (for USD/generic sidearm sites)
        sidearm_staff_items = soup.select('li.sidearm-roster-staff-item')

        # New selector for h3 name format
        h3_staff_names = soup.select('a[href*="/roster/staff/"] h3')

        # Added selector for Clemson's staff table format
        staff_rows = soup.select('tr.person__item')

        # Added selector for Georgia Tech's staff table format
        gt_staff_rows = soup.select('tr:has(td > a[href*="/coaches/"])')

        # New selector for Stanford's staff format
        stanford_staff_links = soup.select('a.table__roster-name[href*="/staff/"]')
        
        # New selector for Virginia Tech staff (using the same class as players, but filtering by URL)
        vt_staff_links = soup.select('a.roster-list-item__title[href*="/staff/"]')

        # New selector for Virginia coaches
        uva_coach_links = soup.select('a[href*="/coach/"]')

        # Process standard staff list formats
        for item in staff_items:
            # Check for the UNC-specific format within the table row
            if 'tr' in item.name and item.has_attr('data-v-7436a2c8'):
                name_tag = item.select_one('td:first-of-type .s-text-regular-bold')
                title_tag = item.select_one('td:last-of-type span')
            else:
                name_tag = item.select_one('.sidearm-roster-coach-name p, .roster-list-item__title')
                title_tag = item.select_one('.sidearm-roster-coach-title span, .roster-list-item__profile-field--position')

            if not name_tag:
                continue

            name = name_tag.get_text(" ", strip=True)
            name = remove_credentials(name)
            title = title_tag.get_text(" ", strip=True) if title_tag else "Staff"

            if "bio" in name.lower() or "view" in name.lower():
                continue

            staff_dict[normalize(name)] = {"name": name, "title": title}

        # Process the newly added Sidearm Sports staff selector
        for item in sidearm_staff_items:
            # Attempt to find name in common Sidearm staff elements
            name_tag = item.select_one('.sidearm-roster-staff-name p, .sidearm-roster-staff-name')
            title_tag = item.select_one('.sidearm-roster-staff-title span, .sidearm-roster-staff-title')
            
            if not name_tag:
                continue
            
            name = name_tag.get_text(" ", strip=True)
            name = remove_credentials(name)
            title = title_tag.get_text(" ", strip=True) if title_tag else "Staff"
            
            staff_dict[normalize(name)] = {"name": name, "title": title}
        
        # Process new h3 name format
        for name_h3 in h3_staff_names:
            name = name_h3.get_text(" ", strip=True)
            name = remove_credentials(name)
            staff_dict[normalize(name)] = {"name": name, "title": "Staff"}

        # Process the new Clemson staff table format
        for row in staff_rows:
            name_tag = row.select_one('td:first-of-type a')
            title_tag = row.select_one('td:nth-of-type(2)')
            if name_tag and title_tag:
                name = name_tag.get_text(" ", strip=True)
                name = remove_credentials(name)
                title = title_tag.get_text(" ", strip=True)
                staff_dict[normalize(name)] = {"name": name, "title": title}
        
        # Process the new Georgia Tech staff table format
        for row in gt_staff_rows:
            name_tag = row.select_one('td > a[href*="/coaches/"]')
            title_tag = name_tag.parent.find_next_sibling('td') if name_tag else None
            if name_tag and title_tag:
                name = name_tag.get_text(" ", strip=True)
                name = remove_credentials(name)
                title = title_tag.get_text(" ", strip=True)
                staff_dict[normalize(name)] = {"name": name, "title": title}

        # Process the new Stanford staff format
        for link in stanford_staff_links:
            name_span = link.select_one('span')
            if name_span:
                name = name_span.get_text(" ", strip=True)
                name = remove_credentials(name)
                staff_dict[normalize(name)] = {"name": name, "title": "Staff"}

        # New selector for Syracuse staff
        syracuse_staff_links = soup.select('div[data-test-id="s-person-details__personal-single-line"] a[href*="/roster/staff/"]')
        for link in syracuse_staff_links:
            name_tag = link.select_one('h3')
            if name_tag:
                name = name_tag.get_text(" ", strip=True)
                name = remove_credentials(name)
                staff_dict[normalize(name)] = {"name": name, "title": "Staff"}
        
        # Process new Virginia Tech staff format
        for link in vt_staff_links:
            name = link.get_text(" ", strip=True)
            name = remove_credentials(name)
            staff_dict[normalize(name)] = {"name": name, "title": "Staff"}

        # New selector for Virginia coaches
        for link in uva_coach_links:
            name = link.get_text(" ", strip=True)
            name = remove_credentials(name)
            staff_dict[normalize(name)] = {"name": name, "title": "Coach"}

        # Additional check for UNC format where coaches are listed in a separate table
        coach_names = soup.select('a[href*="/coaches/"] span.s-text-regular-bold')
        for name_span in coach_names:
            name = name_span.get_text(" ", strip=True)
            name = remove_credentials(name)
            staff_dict[normalize(name)] = {"name": name, "title": "Coach"}

        return staff_dict

    except Exception as e:
        # Improved error logging for debugging
        st.error(f"Error scraping staff names: {e}")
        return {}


def generate_expected_filenames(player_keys, school_prefix):
    """
    Only generate expected filenames for players (ignore staff).
    """
    expected_files = []
    for normalized_name, original_name in player_keys.items():
        parts = original_name.split(" ")
        if len(parts) >= 2:
            first = parts[0].lower()
            last = parts[-1].lower()
            expected_filename = f"{school_prefix}_{last}_{first}.png"
            expected_files.append(expected_filename)
    return expected_files


def find_missing_players(parsed_files, player_keys, staff_dict, school_prefix):
    existing_player_files = set()
    for entry in parsed_files:
        if not entry.get("format_valid", False):
            continue
        normalized_name = normalize(f"{entry.get('first','')} {entry.get('last','')}")
        if normalized_name not in staff_dict:
            existing_player_files.add(normalized_name)

    missing_players = []

    for normalized_name, roster_name in player_keys.items():
        if normalized_name in staff_dict:
            continue

        if roster_name.lower() in ["full bio", "view full bio"]:
            continue

        if normalized_name not in existing_player_files:
            
            # --- START CORRECTED NAME SPLIT LOGIC ---
            
            # Use the normalized_name (e.g., "kalvyn crummie") for splitting
            parts = [p.strip() for p in normalized_name.split() if p.strip()]
            
            first = ""
            last = ""
            
            if len(parts) >= 2:
                # Two or more names: Last is the final word, First is everything before it.
                # Example: ['kalvyn', 'crummie'] -> first='kalvyn', last='crummie'
                # Example: ['d'arious', 't', 'reed'] -> first='d'arious t', last='reed'
                last_raw = parts[-1]
                first_raw = ' '.join(parts[:-1])
                
            elif len(parts) == 1:
                # Single-name case: Use it as the 'first' name.
                # Example: ['tymoss'] -> first='tymoss', last=''
                first_raw = parts[0]
                last_raw = ""
            else:
                continue # Skip empty names

            # Clean up parts for the FILENAME (remove ' and - from the suggested filename components)
            first = re.sub(r"['\-]", "", first_raw)
            last = re.sub(r"['\-]", "", last_raw)
            
            # --- END CORRECTED NAME SPLIT LOGIC ---

            missing_players.append({
                "filename": None,
                "first": first, 
                "last": last,
                "status": "⚠️ Missing"
            })

    return missing_players

# --- Google Drive folder helpers (no API) ---

def _extract_drive_folder_id(url: str) -> str | None:
    """
    Extracts the folder ID from a Google Drive folder URL.
    """
    try:
        parsed = urlparse(url)
        # path-based
        m = re.search(r"/folders/([a-zA-Z0-9_-]+)", parsed.path)
        if m:
            return m.group(1)
        # query-based (?id=)
        qs = parse_qs(parsed.query or "")
        if "id" in qs and len(qs["id"]) > 0:
            return qs["id"][0]
    except:
        pass
    return None

def get_drive_folder_png_filenames(folder_url: str) -> list[str]:
    """
    Fetch PNG filenames from a PUBLIC Google Drive folder without Google API.
    Uses the 'embeddedfolderview' endpoint, which returns parseable HTML.
    """
    folder_id = _extract_drive_folder_id(folder_url)
    if not folder_id:
        st.warning("Could not recognize a Google Drive folder ID from the URL.")
        return []

    # Try embedded folder view first (most reliable for scraping)
    candidates = [
        f"https://drive.google.com/embeddedfolderview?id={folder_id}#list",
        f"https://drive.google.com/embeddedfolderview?id={folder_id}#grid",
        # Fallbacks (may not contain names, but try anyway)
        f"https://drive.google.com/drive/folders/{folder_id}",
        f"https://drive.google.com/drive/u/0/folders/{folder_id}",
    ]

    for url in candidates:
        try:
            # **FIX:** Use headers here too, just in case Drive is sensitive
            r = requests.get(url, timeout=30, headers=COMMON_HEADERS, verify=False)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")

            # Strategy: collect all visible text nodes and filter for *.png
            texts = [t.strip() for t in soup.stripped_strings if t.strip()]
            pngs = [t for t in texts if t.lower().endswith(".png")]

            # Deduplicate while preserving order
            seen = set()
            out = []
            for x in pngs:
                if x not in seen:
                    seen.add(x)
                    out.append(x)

            if out:
                return out
        except Exception:
            continue

    # If nothing found, give a hint
    st.info(
        "No .png names found. Make sure the folder is set to 'Anyone with the link' (Viewer). "
        "Then reload and try again."
    )
    return []

# --- Comparison logic (Unchanged) ---

def check_mismatches_and_missing(parsed_files, player_keys, nickname_keys, staff_dict, school_prefix):
    data = []
    # Inside check_mismatches_and_missing
    parsed_files = [f for f in parsed_files if f.get("format_valid", False)]

    # --- Step 1: existing files ---
    for entry in parsed_files:
        filename = entry.get("filename")
        fmt_ok = entry.get("format_valid", False)
        raw_last = entry.get("last") or ""
        raw_first = entry.get("first") or ""
        suggestion = None
        roster_name = None

        if not fmt_ok:
            status = entry.get("format_msg", "Invalid filename format")
        else:
            school = entry["school"]
            normalized_filename_name = normalize(f"{raw_first} {raw_last}")

            # Staff first - BUT check if this is actually a player with a suffix
            if normalized_filename_name in staff_dict:
                # Check if there's also a player with the same normalized name
                if normalized_filename_name in player_keys:
                    # Both exist - this is likely a Jr/Sr situation
                    # Check the original name from roster to see if it has a suffix
                    original_roster_name = player_keys[normalized_filename_name]
                    
                    # If the roster name has Jr/Sr/II/III/IV/V, it's the player, not staff
                    if re.search(r'\b(jr|sr|ii|iii|iv|v)\b', original_roster_name.lower()):
                        status = "✅"
                        roster_name = player_keys[normalized_filename_name]
                    else:
                        # No suffix in roster name, so it's actually the staff member
                        staff_info = staff_dict[normalized_filename_name]
                        status = f"❌ Not a Player ({staff_info.get('title','Staff')})"
                else:
                    # Only in staff dict, not in players
                    staff_info = staff_dict[normalized_filename_name]
                    status = f"❌ Not a Player ({staff_info.get('title','Staff')})"

            elif school != school_prefix.lower():
                status = "❌ School prefix mismatch"

            elif normalized_filename_name in player_keys:
                status = "✅"
                roster_name = player_keys[normalized_filename_name]
                
                # Check if filename incorrectly includes Jr/Sr/II/III/IV/V as standalone suffixes
                # These suffixes should NOT be in the filename
                first_lower = raw_first.lower()
                last_lower = raw_last.lower()
                
                # Check if suffix exists as a complete word (not part of another word)
                if re.search(r'\b(jr|sr|ii|iii|iv|v)\b', first_lower) or re.search(r'\b(jr|sr|ii|iii|iv|v)\b', last_lower):
                    status = "❌ Filename should not include Jr/Sr/II/III/IV/V suffix"

            elif normalized_filename_name in nickname_keys:
                original_roster_name = nickname_keys[normalized_filename_name]
                match = re.search(r'(\S+)\s+["“”‘’]', original_roster_name)
                if match:
                    first_name = normalize(match.group(1))
                    status = "❌ Nickname used instead of First Name"
                    suggestion = f"{school_prefix}_{raw_last}_{first_name}.png"
                    roster_name = original_roster_name
                else:
                    status = "❌ Nickname used instead of First Name"
                    roster_name = original_roster_name
            else:
                status = "❌ Name not in roster"

        data.append({
            "filename": filename,
            "first": raw_first,
            "last": raw_last,
            "status": status
        })

    # --- Step 2: missing players (ignore staff) ---
    missing_players = find_missing_players(parsed_files, player_keys, staff_dict, school_prefix)
    data.extend(missing_players)

    return pd.DataFrame(data)

# --- Helper function for name validation ---

def is_valid_player_name(name: str) -> bool:
    """
    Check if a name is valid for a player (has at least first and last name).
    Returns False for single-word names or obviously invalid entries.
    """
    if not name or len(name.strip()) == 0:
        return False
    
    # Split by spaces
    parts = name.strip().split()
    
    # Must have at least 2 words (first and last name)
    if len(parts) < 2:
        return False
    
    # Each part should have at least 1 character
    for part in parts:
        clean_part = part.replace('.', '')
        if len(clean_part) == 0:
            return False
    
    return True

# --- Batch Processing Functions ---

def check_single_school(school_name, school_prefix, roster_url, drive_folder_url):
    """
    Check a single school's roster and return results.
    Returns a tuple: (success, result_df, player_count, staff_count, error_message)
    """
    try:
        # Get image files from Google Drive
        image_files = get_drive_folder_png_filenames(drive_folder_url)
        
        if not image_files:
            return (False, None, 0, 0, "No images found in Drive folder")
        
        # Parse filenames
        parsed_files = parse_filenames(image_files)
        
        # Scrape players and staff
        player_keys, nickname_keys = scrape_player_names(roster_url)
        staff_dict = scrape_staff_names(roster_url)
        
        # Remove staff accidentally scraped as players
        for staff_name in list(staff_dict.keys()):
            if staff_name in player_keys:
                player_original_name = player_keys[staff_name]
                has_suffix = bool(re.search(r'\b(jr|sr|ii|iii|iv|v)\b', player_original_name.lower()))
                if not has_suffix:
                    del player_keys[staff_name]
        
        # Remove non-player entries using invalid keywords
        invalid_keywords = [
            "coach", "staff", "jersey", "number", "manager", "director",
            "head coach", "assistant", "trainer", "operations", "headshot", "print",
            "roster", "search", "jump", "video", "feb", "mar", "stats", "baseball",
            "donate", "coaches", "camps"
        ]
        for key in list(player_keys.keys()):
            if contains_invalid_word(player_keys[key], invalid_keywords):
                del player_keys[key]
            if key in nickname_keys:
                del nickname_keys[key]
        
        if not player_keys:
            return (False, None, 0, 0, "No players detected from roster page")
        
        # Check mismatches and missing
        df = check_mismatches_and_missing(
            parsed_files, player_keys, nickname_keys, staff_dict, school_prefix
        )
        
        # Add school name column
        df.insert(0, 'school', school_name)
        
        return (True, df, len(player_keys), len(staff_dict), None)
        
    except Exception as e:
        return (False, None, 0, 0, str(e))

# --- Streamlit UI (main script) ---

st.title("School Roster Photo Name Checker")

# Add mode selection
mode = st.radio("Select Mode:", ["Single School Check", "Batch School Check"])

if mode == "Single School Check":
    # --- SINGLE SCHOOL MODE ---
    st.subheader("Single School Check")
    
    source = st.radio("Where are the images stored?", ["Local folder", "Google Drive folder"])

    image_files: list[str] = []

    if source == "Local folder":
        folder_path = st.text_input("Paste the path to your image folder here:")
        if folder_path and os.path.exists(folder_path):
            image_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".png")]
            st.caption(f"Found {len(image_files)} .png files locally.")
        elif folder_path:
            st.error("Folder path does not exist.")

    elif source == "Google Drive folder":
        drive_folder_url = st.text_input("Paste the PUBLIC Google Drive folder URL here:")
        if drive_folder_url:
            image_files = get_drive_folder_png_filenames(drive_folder_url)
            if image_files:
                st.success(f"Found {len(image_files)} .png files in the Drive folder.")
                with st.expander("Show detected filenames"):
                    st.write(image_files)

    school_prefix = st.text_input("Enter the school prefix (e.g., cal, oregon):")
    school_url = st.text_input("Paste the school roster URL here:")

    if st.button("Check Files"):
        if not image_files:
            st.error("No image files detected yet.")
        elif not school_prefix or not school_url:
            st.error("Please fill in both the school prefix and the roster URL.")
        else:
            parsed_files = parse_filenames(image_files)
            
            # --- Scrape players and staff ---
            player_keys, nickname_keys = scrape_player_names(school_url)
            staff_dict = scrape_staff_names(school_url)
            
            # --- Remove staff accidentally scraped as players ---
            for staff_name in list(staff_dict.keys()):
                if staff_name in player_keys:
                    player_original_name = player_keys[staff_name]
                    has_suffix = bool(re.search(r'\b(jr|sr|ii|iii|iv|v)\b', player_original_name.lower()))
                    if not has_suffix:
                        del player_keys[staff_name]
            
            # --- Remove non-player entries using invalid keywords ---
            invalid_keywords = [
                "coach", "staff", "jersey", "number", "manager", "director",
                "head coach", "assistant", "trainer", "operations", "headshot", "print",
                "roster", "search", "jump", "video", "feb", "mar", "stats", "baseball",
                "donate", "coaches", "camps"
            ]
            for key in list(player_keys.keys()):
                if contains_invalid_word(player_keys[key], invalid_keywords):
                    del player_keys[key]
                if key in nickname_keys:
                    del nickname_keys[key]
            
            if not player_keys:
                st.warning("No players detected from the roster page.")
            else:
                df = check_mismatches_and_missing(
                    parsed_files, player_keys, nickname_keys, staff_dict, school_prefix
                )
                st.subheader("Roster Photo Check")
                st.dataframe(df)

else:
    # --- BATCH SCHOOL MODE (spreadsheet-driven) ---
    st.subheader("Batch School Check")

    # ── Step 1: offer the template download ──────────────────────────────
    st.markdown("### Step 1 — Download the school list template")
    st.write(
        "Fill in your schools (one per row), save the file, "
        "then upload it below. Each row needs: "
        "**Team Abbreviation**, **Roster URL**, and **Google Drive Folder URL**."
    )

    def make_template_bytes() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Schools"
        headers = ["Team Abbreviation", "Roster URL", "Google Drive Folder URL"]
        hdr_fill = PatternFill("solid", start_color="1F4E79")
        hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        examples = [
            ["utah", "https://utahutes.com/sports/football/roster/2025",
             "https://drive.google.com/drive/folders/YOUR_FOLDER_ID"],
            ["arizona", "https://arizonawildcats.com/sports/football/roster/2025",
             "https://drive.google.com/drive/folders/YOUR_FOLDER_ID"],
        ]
        eg_font = Font(name="Times New Roman", color="000000", italic=False, size=12)
        for ri, row_data in enumerate(examples, 2):
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = eg_font
                c.alignment = Alignment(vertical="center", wrap_text=True)
        note = ws.cell(
            row=5, column=1,
            value="<-- Replace example rows above with your own schools. Delete these rows when done."
        )
        note.font = Font(name="Arial", color="C00000", italic=True, size=9)
        ws.merge_cells("A5:C5")
        note.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 62
        ws.column_dimensions["C"].width = 62
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    st.download_button(
        label="📥 Download School List Template (.xlsx)",
        data=make_template_bytes(),
        file_name="roster_schools_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()

    # ── Step 2: upload the filled-in spreadsheet ─────────────────────────
    st.markdown("### Step 2 — Upload your filled-in school list")
    uploaded = st.file_uploader(
        "Upload your Excel school list (.xlsx)",
        type=["xlsx", "xls"],
        help="Must have columns: Team Abbreviation, Roster URL, Google Drive Folder URL",
    )

    schools_df = None
    if uploaded:
        try:
            df_raw = pd.read_excel(uploaded, dtype=str)
            df_raw.columns = [c.strip() for c in df_raw.columns]
            rename_map = {}
            for col in df_raw.columns:
                lower = col.lower()
                if "abbreviation" in lower or "prefix" in lower or lower in ("abbr", "team"):
                    rename_map[col] = "Team Abbreviation"
                elif "roster" in lower:
                    rename_map[col] = "Roster URL"
                elif "drive" in lower or "folder" in lower:
                    rename_map[col] = "Google Drive Folder URL"
            df_raw = df_raw.rename(columns=rename_map)
            required = {"Team Abbreviation", "Roster URL", "Google Drive Folder URL"}
            missing_cols = required - set(df_raw.columns)
            if missing_cols:
                st.error(f"Spreadsheet is missing columns: {', '.join(missing_cols)}")
            else:
                schools_df = df_raw[list(required)].dropna(how="all")
                schools_df = schools_df[
                    schools_df["Roster URL"].str.startswith("http", na=False)
                ].reset_index(drop=True)
                if schools_df.empty:
                    st.error("No valid rows found (Roster URL must start with 'http').")
                    schools_df = None
                else:
                    st.success(f"Loaded **{len(schools_df)} school(s)** from spreadsheet.")
                    with st.expander("Preview school list"):
                        st.dataframe(schools_df, use_container_width=True)
        except Exception as e:
            st.error(f"Could not read spreadsheet: {e}")

    st.divider()

    # ── Step 3: run the batch check ───────────────────────────────────────
    st.markdown("### Step 3 — Run the batch check")

    if schools_df is None:
        st.info("Upload a school list above to enable batch processing.")
    else:
        if st.button(
            f"🔍 Batch Process {len(schools_df)} School(s)",
            type="primary",
            use_container_width=True,
        ):
            from datetime import datetime
            results = []
            progress_bar = st.progress(0)
            status_text = st.empty()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for idx, row in schools_df.iterrows():
                abbr       = str(row["Team Abbreviation"]).strip()
                roster_url = str(row["Roster URL"]).strip()
                drive_url  = str(row["Google Drive Folder URL"]).strip()

                status_text.text(f"Checking {abbr}... ({idx + 1}/{len(schools_df)})")

                success, df, player_count, staff_count, error = check_single_school(
                    abbr, abbr, roster_url, drive_url
                )

                if success:
                    issues = len(df[df["status"] != "✅"])
                    results.append({
                        "school": abbr,
                        "status": "✅ Success",
                        "players": player_count,
                        "staff": staff_count,
                        "issues": issues,
                        "data": df,
                    })
                else:
                    results.append({
                        "school": abbr,
                        "status": f"❌ Error: {error}",
                        "players": 0,
                        "staff": 0,
                        "issues": 0,
                        "data": None,
                    })

                progress_bar.progress((idx + 1) / len(schools_df))

            status_text.text("✅ Batch check complete!")

            # Summary table
            st.subheader("📊 Summary")
            summary_data = [{
                "School": r["school"],
                "Status": r["status"],
                "Players on Roster": r["players"],
                "Issues Found": r["issues"],
            } for r in results]
            st.dataframe(pd.DataFrame(summary_data), use_container_width=True)

            # Detailed results per school
            st.subheader("📋 Detailed Results")
            for r in results:
                if r["data"] is not None:
                    with st.expander(f"{r['school']}  —  {r['issues']} issue(s)"):
                        st.dataframe(r["data"], use_container_width=True)
                        total = len(r["data"])
                        ok    = len(r["data"][r["data"]["status"] == "✅"])
                        st.caption(
                            f"✅ {ok} OK  |  ⚠️ {total - ok} issues  |  📊 {total} total entries"
                        )
                else:
                    with st.expander(f"{r['school']}  —  ERROR"):
                        st.error(r["status"])

            # Downloads
            st.subheader("📥 Download Results")
            col1, col2 = st.columns(2)
            ts_safe = datetime.now().strftime("%Y%m%d_%H%M%S")

            with col1:
                excel_buf = BytesIO()
                with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                    pd.DataFrame(summary_data).to_excel(
                        writer, sheet_name="Summary", index=False
                    )
                    for r in results:
                        if r["data"] is not None:
                            sheet_name = re.sub(r'[:\\\/\?\*\[\]]', "", r["school"])[:31]
                            r["data"].to_excel(writer, sheet_name=sheet_name, index=False)
                excel_buf.seek(0)
                st.download_button(
                    label="📊 Download Excel Report",
                    data=excel_buf,
                    file_name=f"batch_roster_check_{ts_safe}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            with col2:
                # Build plain-text report
                lines = []
                lines.append("=" * 70)
                lines.append("ROSTER PHOTO NAME CHECKER — BATCH REPORT")
                lines.append(f"Generated: {timestamp}")
                lines.append("=" * 70)
                for r in results:
                    lines.append("")
                    lines.append(f"SCHOOL: {r['school'].upper()}")
                    lines.append("-" * 50)
                    if r["data"] is None:
                        lines.append(f"  STATUS: ERROR — {r['status']}")
                        continue
                    df2 = r["data"]
                    total = len(df2)
                    ok    = len(df2[df2["status"] == "✅"])
                    lines.append(f"  Players on roster : {r['players']}")
                    lines.append(f"  Files checked     : {total}")
                    lines.append(f"  OK                : {ok}")
                    lines.append(f"  Issues            : {total - ok}")
                    if total - ok > 0:
                        lines.append("")
                        lines.append("  Issues:")
                        for _, row2 in df2[df2["status"] != "✅"].iterrows():
                            fname = row2.get("filename") or \
                                    f"[{row2.get('first','')}_{row2.get('last','')}]"
                            lines.append(f"    {str(row2['status']):45s}  {fname}")
                lines.append("")
                lines.append("=" * 70)
                lines.append("END OF REPORT")
                text_report = "\n".join(lines)

                st.download_button(
                    label="📄 Download Text Report",
                    data=text_report.encode("utf-8"),
                    file_name=f"batch_roster_check_{ts_safe}.txt",
                    mime="text/plain",
                    use_container_width=True,
                )
